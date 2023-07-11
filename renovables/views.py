from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .forms import ProyectoRenovableForm
from .forms import MontoMMUSDForm
from .forms import MWIncorporadaForm
from .models import ProyectoRenovable
from .models import Monto_PID_MMUSD
from .models import MW_Incorporadas_por_ano
from openpyxl import Workbook

# Create your views here.
@login_required
def proyectos_renovables(request):
    #Variables de busqueda
    query = request.GET.get('q')
    tecnologia = request.GET.get('tecnologia')
    proyectos = ProyectoRenovable.objects.prefetch_related(
        'monto_pid_mmusd_set', 'mw_incorporadas_por_ano_set').all()
    
    #Filtros y barra de busqueda
    if query:
        proyectos = proyectos.filter(nombre_del_proyecto__icontains=query)

    if tecnologia:
        proyectos = proyectos.filter(tecnologia=tecnologia)

    return render(request, 'proyectos_renovables.html', {
        'proyectos': proyectos
    })


@login_required
def crear_proyecto_renovable(request):
    if request.method == 'GET':
        return render(request, 'crear_proyecto_renovable.html', {
            'form': ProyectoRenovableForm
        })
    else:
        try:
            form = ProyectoRenovableForm(request.POST)
            new_proyecto = form.save(commit=False)
            new_proyecto.save()
            return redirect('proyectos_renovables')
        except ValueError:
            return render(request, 'crear_proyecto_renovable.html', {
                'form': ProyectoRenovableForm,
                'error': 'Porfavor, ingresa valores validos'
            })


@login_required
def proyecto_renovable_detalle(request, ProyectoRenovable_id):
    if request.method == 'GET':
        proyecto = get_object_or_404(
            ProyectoRenovable, pk=ProyectoRenovable_id)
        form = ProyectoRenovableForm(instance=proyecto)
        return render(request, 'proyecto_renovable_detalle.html', {
            'proyecto': proyecto,
            'form': form
        })
    else:
        try:
            proyecto = get_object_or_404(
                ProyectoRenovable, pk=ProyectoRenovable_id)
            form = ProyectoRenovableForm(request.POST, instance=proyecto)
            form.save()
            return redirect('proyectos_renovables')
        except ValueError:
            return render(request, 'proyecto_renovable_detalle.html', {
                'proyecto': proyecto,
                'form': form,
                'error': 'Error al actualizar registro'
            })


@login_required
def borrar_proyecto_renovable(request, ProyectoRenovable_id):
    proyecto = get_object_or_404(ProyectoRenovable, pk=ProyectoRenovable_id)
    if request.method == 'POST':
        proyecto.delete()
        return redirect('proyectos_renovables')


@login_required
def excel_renovables(request):
    proyectos = ProyectoRenovable.objects.all()

    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active

    # Escribir los encabezados de las columnas
    columnas = ['Nombre', 'Tecnologia', 'Estado', 'Tipo', 'Esquema de negocio', 'Departamento', 'Ecopetrol/Filial',
                'Vicep/Filial', 'Activo', 'Capacidad Instalada MW/MWp DC', 'Reduccion CO2 (ktCO2/Año)',
                'Inversion Ecopetrol', 'Riesgo No Materializacion', 'Observaciones', 'Lider iniciativa',
                'Monto MMUSD', 'Año Monto', 'Monto MW', 'Año MW']

    for col_num, columna in enumerate(columnas, 1):
        ws.cell(row=1, column=col_num, value=columna)

    # Escribir los datos de los proyectos
    for row_num, proyecto in enumerate(proyectos, 2):
        ws.cell(row=row_num, column=1, value=proyecto.nombre_del_proyecto)
        ws.cell(row=row_num, column=2, value=proyecto.tecnologia)
        ws.cell(row=row_num, column=3, value=proyecto.estado)
        ws.cell(row=row_num, column=4, value=proyecto.tipo)
        ws.cell(row=row_num, column=5, value=proyecto.esquema_de_negocio)
        ws.cell(row=row_num, column=6, value=proyecto.departamento)
        ws.cell(row=row_num, column=7, value=proyecto.ecopetrol_filial)
        ws.cell(row=row_num, column=8, value=proyecto.vicep_filial)
        ws.cell(row=row_num, column=9, value=proyecto.activo)
        ws.cell(row=row_num, column=10,
                value=proyecto.capacidad_instalada_MW_MWp_DC)
        ws.cell(row=row_num, column=11, value=proyecto.reduccion_co2_ktCO2_año)
        ws.cell(row=row_num, column=12, value=proyecto.inversion_ecopetrol)
        ws.cell(row=row_num, column=13, value=proyecto.riesgo_materializacion)
        ws.cell(row=row_num, column=14, value=proyecto.observaciones)
        ws.cell(row=row_num, column=15, value=proyecto.lider_iniciativa)

        # Obtener los datos de la tabla Monto_PID_MMUSD del proyecto actual
        montos = proyecto.monto_pid_mmusd_set.all()
        montos_mmusd = [monto.monto for monto in montos]
        montos_ano = [monto.año for monto in montos]

        # Escribir los datos de Monto_PID_MMUSD
        if montos_mmusd:
            ws.cell(row=row_num, column=16, value=', '.join(str(m)
                    for m in montos_mmusd))
            ws.cell(row=row_num, column=17, value=', '.join(str(a)
                    for a in montos_ano))
        else:
            ws.cell(row=row_num, column=16, value='')
            ws.cell(row=row_num, column=17, value='')

        # Obtener los datos de la tabla MW_Incorporadas_por_ano del proyecto actual
        mws = proyecto.mw_incorporadas_por_ano_set.all()
        mws_mw = [mw.montomw for mw in mws]
        mws_ano = [mw.año for mw in mws]

        # Escribir los datos de MW_Incorporadas_por_ano
        if mws_mw:
            ws.cell(row=row_num, column=18, value=', '.join(str(m)
                    for m in mws_mw))
            ws.cell(row=row_num, column=19, value=', '.join(str(a)
                    for a in mws_ano))
        else:
            ws.cell(row=row_num, column=18, value='')
            ws.cell(row=row_num, column=19, value='')

    # Generar la respuesta del archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proyectos_renovables.xlsx'
    wb.save(response)

    return response

# Rutas para Monto MMUSD


@login_required
def montos_mmusd(request):
    query = request.GET.get('q')
    montos = Monto_PID_MMUSD.objects.all()

    if query:
        montos = montos.filter(proyecto__nombre_del_proyecto__icontains=query)

    return render(request, 'montos.html', {
        'montos': montos
    })


@login_required
def crear_monto_mmusd(request):
    if request.method == 'GET':
        return render(request, 'crear_monto.html', {
            'form': MontoMMUSDForm
        })
    else:
        try:
            form = MontoMMUSDForm(request.POST)
            new_monto = form.save(commit=False)
            new_monto.save()
            return redirect('renovables_montos')
        except ValueError:
            # print(new_monto)
            return render(request, 'crear_monto.html', {
                'form': MontoMMUSDForm,
                'error': 'Porfavor, ingresa valores validos'
            })


@login_required
def monto_detalle(request, monto_id):
    if request.method == 'GET':
        monto = get_object_or_404(
            Monto_PID_MMUSD, pk=monto_id)
        form = MontoMMUSDForm(instance=monto)
        return render(request, 'monto_detalle.html', {
            'monto': monto,
            'form': form
        })
    else:
        try:
            monto = get_object_or_404(
                Monto_PID_MMUSD, pk=monto_id)
            form = MontoMMUSDForm(request.POST, instance=monto)
            form.save()
            return redirect('renovables_montos')
        except ValueError:
            return render(request, 'monto_detalle.html', {
                'monto': monto,
                'form': form,
                'error': 'Error al actualizar registro'
            })


@login_required
def borrar_monto(request, monto_id):
    monto = get_object_or_404(Monto_PID_MMUSD, pk=monto_id)
    if request.method == 'POST':
        monto.delete()
        return redirect('renovables_montos')


@login_required
def excel_renovables_montos(request):
    # Reemplaza TusProyectosModel con tu modelo real
    montos = Monto_PID_MMUSD.objects.all()

    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active

    # Escribir los encabezados de las columnas
    columnas = ['Proyecto', 'Año', 'Monto']

    for col_num, columna in enumerate(columnas, 1):
        ws.cell(row=1, column=col_num, value=columna)

    # Escribir los datos de los proyectos
    for row_num, monto in enumerate(montos, 2):
        ws.cell(row=row_num, column=1, value=monto.proyecto.nombre_del_proyecto)
        ws.cell(row=row_num, column=2, value=monto.año)
        ws.cell(row=row_num, column=3, value=monto.monto)

    # Generar la respuesta del archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proyectos renovables Montos MMUSD.xlsx'
    wb.save(response)

    return response

# Rutas para MW Incorporadas


@login_required
def montos_mw(request):
    query = request.GET.get('q')
    montosmw = MW_Incorporadas_por_ano.objects.all()

    if query:
        montosmw = montosmw.filter(proyecto__nombre_del_proyecto__icontains=query)

    return render(request, 'montosmw.html', {
        'montos': montosmw
    })


@login_required
def crear_monto_mw(request):
    if request.method == 'GET':
        return render(request, 'crear_montomw.html', {
            'form': MWIncorporadaForm
        })
    else:
        try:
            form = MWIncorporadaForm(request.POST)
            new_monto = form.save(commit=False)
            new_monto.save()
            return redirect('renovables_montosmw')
        except ValueError:
            return render(request, 'crear_montomw.html', {
                'form': MWIncorporadaForm,
                'error': 'Porfavor, ingresa valores validos'
            })


@login_required
def montomw_detalle(request, monto_id):
    if request.method == 'GET':
        monto = get_object_or_404(
            MW_Incorporadas_por_ano, pk=monto_id)
        form = MWIncorporadaForm(instance=monto)
        return render(request, 'montomw_detalle.html', {
            'monto': monto,
            'form': form
        })
    else:
        try:
            monto = get_object_or_404(
                MW_Incorporadas_por_ano, pk=monto_id)
            form = MWIncorporadaForm(request.POST, instance=monto)
            form.save()
            return redirect('renovables_montosmw')
        except ValueError:
            return render(request, 'montomw_detalle.html', {
                'monto': monto,
                'form': form,
                'error': 'Error al actualizar registro'
            })


@login_required
def borrar_montomw(request, monto_id):
    monto = get_object_or_404(MW_Incorporadas_por_ano, pk=monto_id)
    if request.method == 'POST':
        monto.delete()
        return redirect('renovables_montosmw')


@login_required
def excel_renovables_montosmw(request):
    # Reemplaza TusProyectosModel con tu modelo real
    montos = MW_Incorporadas_por_ano.objects.all()

    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active

    # Escribir los encabezados de las columnas
    columnas = ['Proyecto', 'Año', 'Monto(MW)']

    for col_num, columna in enumerate(columnas, 1):
        ws.cell(row=1, column=col_num, value=columna)

    # Escribir los datos de los proyectos
    for row_num, monto in enumerate(montos, 2):
        ws.cell(row=row_num, column=1, value=monto.proyecto.nombre_del_proyecto)
        ws.cell(row=row_num, column=2, value=monto.año)
        ws.cell(row=row_num, column=3, value=monto.montomw)

    # Generar la respuesta del archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proyectos renovables MW Incorporadas por año.xlsx'
    wb.save(response)

    return response
