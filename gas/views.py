from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import ProyectoGas
from .models import Inversiones_MUSD
from .forms import ProyectoGasForm
from .forms import InversionesMUSDForm
from openpyxl import Workbook

# Create your views here.


@login_required
def proyectos_gas(request):
    query = request.GET.get('q')
    segmento = request.GET.get('segmento')
    proyectos = ProyectoGas.objects.prefetch_related(
        'inversiones_musd_set').all()
    proyectos = ProyectoGas.objects.all()

    # Filtros y barra de busqueda
    if query:
        proyectos = proyectos.filter(nombre_del_proyecto__icontains=query)

    if segmento:
        proyectos = proyectos.filter(departamento=segmento)

    return render(request, 'proyectos_gas.html', {
        'proyectos': proyectos
    })


@login_required
def crear_proyecto_gas(request):
    if request.method == 'GET':
        return render(request, 'crear_proyecto_gas.html', {
            'form': ProyectoGasForm
        })
    else:
        try:
            form = ProyectoGasForm(request.POST)
            new_proyecto = form.save(commit=False)
            new_proyecto.save()
            return redirect('proyectos_gas')
        except ValueError:
            print(form.errors)
            return render(request, 'crear_proyecto_gas.html', {
                'form': ProyectoGasForm,
                'error': str('Porfavor, ingresa valores validos' + form.errors)
            })


@login_required
def proyecto_gas_detalle(request, ProyectoGas_id):
    if request.method == 'GET':
        proyecto = get_object_or_404(
            ProyectoGas, pk=ProyectoGas_id)
        form = ProyectoGasForm(instance=proyecto)
        return render(request, 'proyecto_gas_detalle.html', {
            'proyecto': proyecto,
            'form': form
        })
    else:
        try:
            proyecto = get_object_or_404(
                ProyectoGas, pk=ProyectoGas_id)
            form = ProyectoGasForm(request.POST, instance=proyecto)
            form.save()
            return redirect('proyectos_gas')
        except ValueError:
            # print(form.errors)
            return render(request, 'proyecto_gas_detalle.html', {
                'proyecto': proyecto,
                'form': form,
                'error': 'Error al actualizar registro, recuereda revisar las fechas con formato yyyy-mm-dd'
            })


@login_required
def borrar_proyecto_gas(request, ProyectoGas_id):
    proyecto = get_object_or_404(ProyectoGas, pk=ProyectoGas_id)
    if request.method == 'POST':
        proyecto.delete()
        return redirect('proyectos_gas')


@login_required
def excel_gas(request):
    proyectos = ProyectoGas.objects.all()

    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active

    # Escribir los encabezados de las columnas
    columnas = ['Matricula Digital', 'Nombre del Proyecto', 'Escenario', 'Empresa', 'Segmento', 'Subsegmento', 'Unidad de Negocios Regional',
                'Gerencia', 'Activo', 'Departamento', 'Categoria',
                'Subcategoria', 'Fase Actual', 'Riesgo', 'Dimension',
                'Ejecutor', 'Variante', 'Proxima Toma de Decision', 'Fecha FID', 'Gas', 'Seguimiento', 'Palanca Tecnologica', 'Cuenca Venture', 'Hidrocarburo', 'Recobro', 'CapEx Transicion',
                'VPN (MUSD)', '(E) VPN (MUSD)', 'EFI (veces)', 'TIR',
                'Payback Meses', 'BE Crudo', 'BE Gas', 'CapEx Unitario',
                'First Oil Gas', 'Volumenes 1P', 'Volumenes 2P', 'Volumenes 3P', 'Recursos Contigentes', 'Rec. Desc. Delimitar (Verificar)', 'Recursos Prospectivos', 'Inversion MUSD (Mes)', 'Inversion MUSD (Año)', 'Inversion MUSD']

    for col_num, columna in enumerate(columnas, 1):
        ws.cell(row=1, column=col_num, value=columna)

    # Escribir los datos de los proyectos
    for row_num, proyecto in enumerate(proyectos, 2):
        ws.cell(row=row_num, column=1, value=proyecto.matricula_digital)
        ws.cell(row=row_num, column=2, value=proyecto.nombre_del_proyecto)
        ws.cell(row=row_num, column=3, value=proyecto.escenario)
        ws.cell(row=row_num, column=4, value=proyecto.empresa)
        ws.cell(row=row_num, column=5, value=proyecto.segmento)
        ws.cell(row=row_num, column=6, value=proyecto.subsegmento)
        ws.cell(row=row_num, column=7,
                value=proyecto.unidad_de_negocios_regional)
        ws.cell(row=row_num, column=8, value=proyecto.gerencia)
        ws.cell(row=row_num, column=9, value=proyecto.activo)
        ws.cell(row=row_num, column=10,
                value=proyecto.departamento)
        ws.cell(row=row_num, column=11, value=proyecto.categoria)
        ws.cell(row=row_num, column=12, value=proyecto.subcategoria)
        ws.cell(row=row_num, column=13, value=proyecto.fase_actual)
        ws.cell(row=row_num, column=14, value=proyecto.riesgo)
        ws.cell(row=row_num, column=15, value=proyecto.dimension)
        ws.cell(row=row_num, column=16, value=proyecto.ejecutor)
        ws.cell(row=row_num, column=17, value=proyecto.variante)
        ws.cell(row=row_num, column=18, value=proyecto.proxima_toma_de_decision)
        ws.cell(row=row_num, column=19, value=proyecto.fecha_fid)
        ws.cell(row=row_num, column=20, value=proyecto.gas)
        ws.cell(row=row_num, column=21, value=proyecto.seguimiento)
        ws.cell(row=row_num, column=22, value=proyecto.palanca_tecnologica)
        ws.cell(row=row_num, column=23, value=proyecto.cuenca_venture)
        ws.cell(row=row_num, column=24, value=proyecto.hidrocarburo)
        ws.cell(row=row_num, column=25, value=proyecto.recobro)
        ws.cell(row=row_num, column=26, value=proyecto.capex_transicion)
        ws.cell(row=row_num, column=27, value=proyecto.vpn_musd)
        ws.cell(row=row_num, column=28, value=proyecto.e_vpn_musd)
        ws.cell(row=row_num, column=29, value=proyecto.efi_veces)
        ws.cell(row=row_num, column=30, value=proyecto.tir)
        ws.cell(row=row_num, column=31, value=proyecto.parback_meses)
        ws.cell(row=row_num, column=32, value=proyecto.be_crudo)
        ws.cell(row=row_num, column=33, value=proyecto.be_gas)
        ws.cell(row=row_num, column=34, value=proyecto.capex_unitario)
        ws.cell(row=row_num, column=35, value=proyecto.first_oil_gas)
        ws.cell(row=row_num, column=36, value=proyecto.volumenes_1p)
        ws.cell(row=row_num, column=37, value=proyecto.volumenes_2p)
        ws.cell(row=row_num, column=38, value=proyecto.volumenes_3p)
        ws.cell(row=row_num, column=39, value=proyecto.recursos_contigentes)
        ws.cell(row=row_num, column=40,
                value=proyecto.rec_desc_delimitar_verificar)
        ws.cell(row=row_num, column=41, value=proyecto.recursos_prospectivos)

        # Obtener los datos de la tabla Monto_PID_MMUSD del proyecto actual
        inversiones = proyecto.inversiones_musd_set.all()
        inversion_mes = [inversion.mes for inversion in inversiones]
        inversion_ano = [inversion.año for inversion in inversiones]
        inversion_monto = [inversion.monto for inversion in inversiones]

        # Escribir los datos de Monto_PID_MMUSD
        if inversion_monto:
            ws.cell(row=row_num, column=42, value=', '.join(str(m)
                    for m in inversion_mes))
            ws.cell(row=row_num, column=43, value=', '.join(str(a)
                    for a in inversion_ano))
            ws.cell(row=row_num, column=44, value=', '.join(str(a)
                    for a in inversion_monto))
        else:
            ws.cell(row=row_num, column=42, value='')
            ws.cell(row=row_num, column=43, value='')
            ws.cell(row=row_num, column=44, value='')

        # Obtener los datos de la tabla MW_Incorporadas_por_ano del proyecto actual
        """ mws = proyecto.mw_incorporadas_por_ano_set.all()
        mws_mw = [mw.montomw for mw in mws]
        mws_ano = [mw.año for mw in mws]

        # Escribir los datos de MW_Incorporadas_por_ano
        if mws_mw:
            ws.cell(row=row_num, column=18, value=', '.join(str(m)
                    for m in mws_mw))
            ws.cell(row=row_num, column=19, value=', '.join(str(a)
                    for a in mws_ano))
        else:
            ws.cell(row=row_num, column=18, value='')
            ws.cell(row=row_num, column=19, value='') """

    # Generar la respuesta del archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proyectos_gas.xlsx'
    wb.save(response)

    return response


@login_required
def inversiones_musd(request):
    query = request.GET.get('q')
    año = request.GET.get('año')
    inversiones = Inversiones_MUSD.objects.all()

    # Filtros y barra de busqueda
    if query:
        inversiones = inversiones.filter(
            proyecto__nombre_del_proyecto__icontains=query)

    if año:
        inversiones = inversiones.filter(año=año)

    return render(request, 'inversiones_musd.html', {
        'inversiones': inversiones
    })


@login_required
def crear_inversion_musd(request):
    if request.method == 'GET':
        return render(request, 'crear_inversion_musd.html', {
            'form': InversionesMUSDForm
        })
    else:
        try:
            form = InversionesMUSDForm(request.POST)
            new_inversion = form.save(commit=False)
            new_inversion.save()
            return redirect('inversiones_musd')
        except ValueError:
            print(form.errors)
            return render(request, 'crear_inversion_musd.html', {
                'form': InversionesMUSDForm,
                'error': 'Porfavor, ingresa valores validos'
            })


def inversion_musd_detalle(request, InversionMusd_id):
    if request.method == 'GET':
        inversion = get_object_or_404(
            Inversiones_MUSD, pk=InversionMusd_id)
        form = InversionesMUSDForm(instance=inversion)
        return render(request, 'inversion_musd_detalle.html', {
            'inversion': inversion,
            'form': form
        })
    else:
        try:
            inversion = get_object_or_404(
                Inversiones_MUSD, pk=InversionMusd_id)
            form = InversionesMUSDForm(request.POST, instance=inversion)
            form.save()
            return redirect('inversiones_musd')
        except ValueError:
            # print(form.errors)
            return render(request, 'inversion_musd_detalle.html', {
                'inversion': inversion,
                'form': form,
                'error': 'Error al actualizar registro'
            })


def borrar_inversion_musd(request, InversionMusd_id):
    inversion = get_object_or_404(Inversiones_MUSD, pk=InversionMusd_id)
    if request.method == 'POST':
        inversion.delete()
        return redirect('inversiones_musd')
